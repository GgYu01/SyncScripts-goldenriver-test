#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
This script automates the process of filling in the Release Note Excel sheet by extracting commit information from multiple Git repositories.
It is designed to run on Linux systems with Python 3.8.10.

Usage:
    - Configure the parameters and variables in the script as needed.
    - Run the script directly: `./release_note_updater.py`

Note:
    - The script is highly modularized and follows professional coding standards.
    - It requires the `openpyxl` and `rich` Python packages for handling Excel files and enhanced console output respectively.
    - Detailed execution logs are generated in the user's home directory.
"""

import os
import sys
import subprocess
import datetime
import logging
from dataclasses import dataclass, field
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.cell_range import CellRange
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn
from rich.logging import RichHandler

# Set up logging configuration with Rich
console = Console()
logging.basicConfig(
    level=logging.INFO,
    format='%(message)s',
    datefmt='[%X]',
    handlers=[RichHandler(console=console)]
)
logger = logging.getLogger('release_note_updater')

# Global constants
HOME_DIR = os.path.expanduser('~')
EXCLUDE_DIRS = {'.repo', '.jiri'}
EXCEL_FILE_PATH = os.path.join(HOME_DIR, 'MT8678-8676_Hypervisor_Release_Note.xlsx')
LOG_FILE_PATH = os.path.join(HOME_DIR, 'release_note_updater.log')

@dataclass
class Config:
    """
    Configuration parameters for the script.
    """
    tester: str = '高宇轩'
    modifier: str = '武阳'
    mtk_owner: str = '金春阳'
    porting_to_other_platform: str = 'No'
    porting_and_testing_done: str = 'No'
    release_to_customer: str = 'No'
    latest_version_identifier: str = ''  # For debugging purposes
    previous_version_identifier: str = ''  # For debugging purposes

@dataclass
class RepositoryInfo:
    """
    Holds information about a Git repository.
    """
    path: str
    module_name: str
    tag_prefix: str
    latest_tag: str = ''
    previous_tag: str = ''
    commits: List[Dict[str, str]] = field(default_factory=list)
    patches: List[str] = field(default_factory=list)

class ReleaseNoteUpdater:
    """
    Main class for updating the Release Note Excel sheet.
    """

    def __init__(self, config: Config):
        self.config = config
        self.repositories = self._initialize_repositories()
        self.zircon_commit_id = ''
        self.garnet_commit_id = ''
        self.setup_logging()

    def setup_logging(self):
        """
        Setup file logging.
        """
        file_handler = logging.FileHandler(LOG_FILE_PATH, mode='w', encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
        file_handler.setFormatter(formatter)
        logging.getLogger().addHandler(file_handler)

    def _initialize_repositories(self) -> List[RepositoryInfo]:
        """
        Initialize repository information based on predefined paths and modules.
        """
        repos = [
            RepositoryInfo(path=os.path.join(HOME_DIR, 'grpower'), module_name='nebula-hyper', tag_prefix='release-spm.mt8678'),
            RepositoryInfo(path=os.path.join(HOME_DIR, 'grt'), module_name='thyp-sdk', tag_prefix='release-spm.mt8678'),
            RepositoryInfo(path=os.path.join(HOME_DIR, 'grt_be'), module_name='thyp-sdk-be', tag_prefix='release-spm.mt8678'),
        ]
        # Recursively find all git repositories under specified directories
        additional_paths = [
            (os.path.join(HOME_DIR, 'grpower', 'workspace', 'nebula'), 'nebula-hyper', 'release-spm.mt8678_mt8676'),
            (os.path.join(HOME_DIR, 'alps'), 'alps', 'release-spm.mt8678'),
            (os.path.join(HOME_DIR, 'yocto'), 'yocto', 'release-spm.mt8678'),
        ]
        for base_path, module_name, tag_prefix in additional_paths:
            for repo_path in self._find_git_repos(base_path):
                repos.append(RepositoryInfo(path=repo_path, module_name=module_name, tag_prefix=tag_prefix))
        logger.info('Initialized repository information.')
        return repos

    def _find_git_repos(self, base_path: str) -> List[str]:
        """
        Recursively find all git repositories under a base path.

        Args:
            base_path (str): The base directory to search.

        Returns:
            List[str]: A list of paths to git repositories.
        """
        git_repos = []
        for root, dirs, files in os.walk(base_path):
            # Skip excluded directories
            dirs[:] = [d for d in dirs if d not in EXCLUDE_DIRS]
            if '.git' in dirs:
                git_repos.append(root)
                # Do not traverse subdirectories of a git repository
                dirs[:] = []
        return git_repos

    def run(self):
        """
        Main execution method.
        """
        logger.info('Starting Release Note Updater...')
        with Progress(SpinnerColumn(), TextColumn("[progress.description]{task.description}")) as progress:
            task = progress.add_task("Processing...", total=None)
            self._get_latest_tags()
            self._collect_commits_and_patches()
            self._get_special_commit_ids()
            self._update_excel()
            progress.update(task, description="Processing complete")
        logger.info('Release Note Updater finished successfully.')

    def _get_latest_tags(self):
        """
        Get the latest and previous tags for each repository.
        """
        # Using grt repository as reference
        grt_repo = next((repo for repo in self.repositories if os.path.basename(repo.path) == 'grt'), None)
        if grt_repo is None:
            logger.error('grt repository not found.')
            sys.exit(1)

        if self.config.latest_version_identifier and self.config.previous_version_identifier:
            # Use debug version identifiers
            grt_repo.latest_tag = self._construct_tag(grt_repo.tag_prefix, self.config.latest_version_identifier)
            grt_repo.previous_tag = self._construct_tag(grt_repo.tag_prefix, self.config.previous_version_identifier)
        else:
            grt_repo.latest_tag, grt_repo.previous_tag = self._get_latest_two_tags(grt_repo.path)
            self.config.latest_version_identifier = self._extract_version_identifier(grt_repo.latest_tag)
            self.config.previous_version_identifier = self._extract_version_identifier(grt_repo.previous_tag)

        # Generate tags for other repositories
        for repo in self.repositories:
            if repo is not grt_repo:
                repo.latest_tag = self._construct_tag(repo.tag_prefix, self.config.latest_version_identifier)
                repo.previous_tag = self._construct_tag(repo.tag_prefix, self.config.previous_version_identifier)

        logger.info('Retrieved latest and previous tags for all repositories.')

    def _get_latest_two_tags(self, repo_path: str) -> Tuple[str, str]:
        """
        Get the latest and previous tags in a repository.

        Args:
            repo_path (str): Path to the Git repository.

        Returns:
            Tuple[str, str]: Latest and previous tags.
        """
        cmd = ['git', '-C', repo_path, 'tag', '--sort=-creatordate']
        result = subprocess.run(cmd, stdout=subprocess.PIPE, text=True)
        tags = result.stdout.strip().split('\n')
        if len(tags) < 2:
            logger.error(f'Not enough tags found in repository: {repo_path}')
            sys.exit(1)
        return tags[0], tags[1]

    def _extract_version_identifier(self, tag: str) -> str:
        """
        Extract the version identifier from a tag.

        Args:
            tag (str): The tag string.

        Returns:
            str: The version identifier.
        """
        parts = tag.split('_')
        if len(parts) >= 3:
            return '_'.join(parts[-3:])
        else:
            logger.error(f'Invalid tag format: {tag}')
            sys.exit(1)

    def _construct_tag(self, prefix: str, version_identifier: str) -> str:
        """
        Construct a tag using the prefix and version identifier.

        Args:
            prefix (str): The tag prefix.
            version_identifier (str): The version identifier.

        Returns:
            str: The constructed tag.
        """
        return f'{prefix}_{version_identifier}'

    def _collect_commits_and_patches(self):
        """
        Collect commits and patches from repositories.
        """
        logger.info('Collecting commits and generating patches...')
        for repo in self.repositories:
            if not os.path.exists(repo.path):
                logger.warning(f'Repository path does not exist: {repo.path}')
                continue
            repo.commits = self._get_commits_between_tags(repo)
            if repo.commits:
                repo.patches = self._generate_patches(repo)
        logger.info('Collected commits and patches from repositories.')

    def _get_commits_between_tags(self, repo: RepositoryInfo) -> List[Dict[str, str]]:
        """
        Get commits between the latest and previous tags.

        Args:
            repo (RepositoryInfo): Repository information.

        Returns:
            List[Dict[str, str]]: List of commits.
        """
        # Check if tags exist in the repository
        existing_tags = self._get_all_tags(repo.path)
        if repo.previous_tag not in existing_tags or repo.latest_tag not in existing_tags:
            logger.warning(f"Tags {repo.previous_tag} or {repo.latest_tag} do not exist in {repo.path}. Skipping repository.")
            return []
        cmd = ['git', '-C', repo.path, 'log', f'{repo.previous_tag}..{repo.latest_tag}', '--pretty=format:%H|%s']
        result = subprocess.run(cmd, stdout=subprocess.PIPE, text=True)
        commits = []
        for line in result.stdout.strip().split('\n'):
            if line:
                commit_hash, commit_message = line.split('|', 1)
                commits.append({'hash': commit_hash, 'message': commit_message.strip()})
        return commits

    def _get_all_tags(self, repo_path: str) -> List[str]:
        """
        Get all tags in a repository.

        Args:
            repo_path (str): Path to the Git repository.

        Returns:
            List[str]: List of tags.
        """
        cmd = ['git', '-C', repo_path, 'tag']
        result = subprocess.run(cmd, stdout=subprocess.PIPE, text=True)
        tags = result.stdout.strip().split('\n')
        return tags

    def _generate_patches(self, repo: RepositoryInfo) -> List[str]:
        """
        Generate patches between the latest and previous tags.

        Args:
            repo (RepositoryInfo): Repository information.

        Returns:
            List[str]: List of patch file paths.
        """
        cmd = ['git', '-C', repo.path, 'format-patch', f'{repo.previous_tag}..{repo.latest_tag}', '-o', repo.path]
        subprocess.run(cmd)
        patches = [os.path.join(repo.path, f) for f in os.listdir(repo.path) if f.endswith('.patch')]
        return patches

    def _get_special_commit_ids(self):
        """
        Get the last commit IDs from specific repositories.
        """
        logger.info('Retrieving special commit IDs...')
        self.zircon_commit_id = self._get_last_commit_id(os.path.join(HOME_DIR, 'grpower', 'workspace', 'nebula', 'zircon'))
        self.garnet_commit_id = self._get_last_commit_id(os.path.join(HOME_DIR, 'grpower', 'workspace', 'nebula', 'garnet'))
        logger.info('Retrieved special commit IDs.')

    def _get_last_commit_id(self, repo_path: str) -> str:
        """
        Get the last commit ID from a repository.

        Args:
            repo_path (str): Path to the repository.

        Returns:
            str: The commit ID.
        """
        if not os.path.exists(repo_path):
            logger.error(f'Repository path does not exist: {repo_path}')
            return ''
        cmd = ['git', '-C', repo_path, 'rev-parse', 'HEAD']
        result = subprocess.run(cmd, stdout=subprocess.PIPE, text=True)
        return result.stdout.strip()

    def _update_excel(self):
        """
        Update the Excel sheet with collected data.
        """
        logger.info('Updating Excel sheet...')
        if not os.path.exists(EXCEL_FILE_PATH):
            logger.error(f'Excel file not found: {EXCEL_FILE_PATH}')
            sys.exit(1)
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

        # Remove unmerging and re-merging of cells
        # Insert new rows from the second row
        new_rows = self._prepare_excel_rows()
        num_new_rows = len(new_rows)
        ws.insert_rows(idx=2, amount=num_new_rows)

        # Fill in data for new rows
        for idx, row_data in enumerate(new_rows):
            for col_num, value in enumerate(row_data, start=1):
                ws.cell(row=2 + idx, column=col_num, value=value)

        wb.save(EXCEL_FILE_PATH)
        logger.info('Excel file updated successfully.')

    def _prepare_excel_rows(self) -> List[List[str]]:
        """
        Prepare the data rows to be inserted into the Excel sheet.

        Returns:
            List[List[str]]: List of rows with data.
        """
        rows = []
        release_version = next((repo.latest_tag for repo in self.repositories if os.path.basename(repo.path) == 'grt'), '')
        submit_info = f'zircon: {self.zircon_commit_id}\ngarnet: {self.garnet_commit_id}'
        current_date = datetime.datetime.now().strftime('%Y-%m-%d')

        for repo in self.repositories:
            for commit in repo.commits:
                row = [''] * 15  # Initialize empty row with 15 columns
                # Column 1: release 版本
                row[0] = release_version
                # Column 2: 功能
                row[1] = commit['message']
                # Column 3: 模块
                row[2] = repo.module_name
                # Column 4: 压缩包/Patch
                patch_files = self._get_patches_for_commit(repo, commit['hash'])
                row[3] = '\n'.join(patch_files)
                # Column 5: 示例代码/文档路径 (left blank)
                # Column 6: 提交信息
                row[5] = submit_info
                # Column 7: 测试负责人 / 修改人 / MTK owner
                row[6] = f"{self.config.tester} / {self.config.modifier} / {self.config.mtk_owner}"
                # Column 8: 提交时间
                row[7] = current_date
                # Column 9: 是否向另外一个平台移植？
                row[8] = self.config.porting_to_other_platform
                # Column 10: 是否已经移植和测试？
                row[9] = self.config.porting_and_testing_done
                # Column 11: 是否Release客户及客户名
                row[10] = self.config.release_to_customer
                # Column 12: change ID / commit ID title (left blank)
                # Column 13: MTK合入日期 (left blank)
                # Column 14: MTK注册情况 (left blank)
                # Column 15: commit信息
                row[14] = commit['hash']
                rows.append(row)
        return rows

    def _get_patches_for_commit(self, repo: RepositoryInfo, commit_hash: str) -> List[str]:
        """
        Get the patch files corresponding to a commit.

        Args:
            repo (RepositoryInfo): Repository information.
            commit_hash (str): Commit hash.

        Returns:
            List[str]: List of relative patch file paths.
        """
        matching_patches = []
        for patch in repo.patches:
            with open(patch, 'r') as f:
                first_line = f.readline()
                if commit_hash in first_line:
                    relative_patch_path = os.path.relpath(patch, '/home/nebula')
                    matching_patches.append(relative_patch_path)
        return matching_patches

if __name__ == '__main__':
    # Initialize configuration with default values or from command-line arguments
    config = Config(
        latest_version_identifier='2024_1107_04',  # Example value for debugging
        previous_version_identifier='2024_1107_03'  # Example value for debugging
    )
    # You can add command-line argument parsing here if needed
    updater = ReleaseNoteUpdater(config)
    updater.run()
